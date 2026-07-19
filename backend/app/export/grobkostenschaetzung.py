"""Schlichte, prüfbare Exporte der BKP-Grobkostenschätzung."""
import io
from datetime import date
from pathlib import Path

import reportlab
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


DUNKEL = "243247"
GRAU = "64748B"
LINIE = "DDE3EA"
HELL = "F7F9FB"
ROT = "B42318"
ROT_HELL = "FEECEB"
AMBER = "A15C07"
AMBER_HELL = "FFF4D6"
BLAU = "175CD3"
BLAU_HELL = "EAF2FF"
GRUEN = "067647"
GRUEN_HELL = "E8F5EE"

_REPORTLAB_FONTS = Path(reportlab.__file__).parent / "fonts"
pdfmetrics.registerFont(TTFont("GkSans", str(_REPORTLAB_FONTS / "Vera.ttf")))
pdfmetrics.registerFont(TTFont("GkSans-Bold", str(_REPORTLAB_FONTS / "VeraBd.ttf")))
pdfmetrics.registerFontFamily("GkSans", normal="GkSans", bold="GkSans-Bold")


def _chf(wert):
    if wert is None:
        return "Keine Angaben"
    return f"CHF {round(wert):,}".replace(",", "'")


def _zahl(wert, stellen=0):
    if wert is None:
        return "-"
    return f"{wert:,.{stellen}f}".replace(",", "'")


def _status_farbe(position):
    if position.get("quelle") == "manuell":
        return BLAU, BLAU_HELL
    status = position.get("status_datenbasis") or ""
    if status == "Keine Angaben":
        return ROT, ROT_HELL
    if status in {"Einzelfall - nicht belastbar", "Einzelfall – nicht belastbar",
                  "Sehr geringe Datengrundlage", "Begrenzte Datengrundlage"}:
        return AMBER, AMBER_HELL
    return GRUEN, GRUEN_HELL


def _alle_positionen(result):
    return [p for g in (result.get("gruppen") or []) for p in (g.get("positionen") or [])]


def _pdf_styles():
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("GkTitle", parent=styles["Title"], fontName="GkSans-Bold",
                                fontSize=21, leading=25, textColor=colors.HexColor(f"#{DUNKEL}"),
                                spaceAfter=5 * mm),
        "h1": ParagraphStyle("GkH1", parent=styles["Heading1"], fontName="GkSans-Bold",
                             fontSize=12, leading=15, textColor=colors.HexColor(f"#{DUNKEL}"),
                             spaceBefore=4 * mm, spaceAfter=2 * mm),
        "body": ParagraphStyle("GkBody", parent=styles["BodyText"], fontName="GkSans",
                               fontSize=8.5, leading=11, textColor=colors.HexColor(f"#{DUNKEL}")),
        "small": ParagraphStyle("GkSmall", parent=styles["BodyText"], fontName="GkSans",
                                fontSize=7, leading=9, textColor=colors.HexColor(f"#{GRAU}")),
        "right": ParagraphStyle("GkRight", parent=styles["BodyText"], fontName="GkSans-Bold",
                                fontSize=8, leading=10, alignment=TA_RIGHT,
                                textColor=colors.HexColor(f"#{DUNKEL}")),
    }


def erzeuge_grobkostenschaetzung_pdf(projekt_name: str, inputs: dict, result: dict,
                                     variante: str) -> bytes:
    """Erzeugt einen kompakten Bericht mit Schätzung und Herkunftsnachweis."""
    buf = io.BytesIO()
    styles = _pdf_styles()

    def fuss(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor(f"#{LINIE}"))
        canvas.line(18 * mm, 14 * mm, 192 * mm, 14 * mm)
        canvas.setFillColor(colors.HexColor(f"#{GRAU}"))
        canvas.setFont("GkSans", 7)
        canvas.drawString(18 * mm, 9 * mm, f"{projekt_name} - Grobkostenschaetzung - {date.today():%d.%m.%Y}")
        canvas.drawRightString(192 * mm, 9 * mm, f"Seite {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=19 * mm,
        title=f"{projekt_name} - Grobkostenschaetzung",
    )
    story = [
        Paragraph("Grobkostenschätzung Heizung", styles["title"]),
        Paragraph(f"<b>{projekt_name}</b> &nbsp;&nbsp; Stand {date.today():%d.%m.%Y} &nbsp;&nbsp; Variante: {variante.title()}", styles["body"]),
        Spacer(1, 5 * mm),
    ]

    grundlagen = [
        ["Nutzung", inputs.get("nutzung") or "-", "Projektart", inputs.get("projektart") or "-"],
        ["Wärmeerzeuger", ", ".join(inputs.get("waermeerzeuger") or []) or "-",
         "Wärmeabgabe", ", ".join(inputs.get("waermeabgabe") or []) or "-"],
        ["EBF", f"{_zahl(inputs.get('ebf_m2'))} m²", "Heizleistung", f"{_zahl(inputs.get('leistung_kw'), 1)} kW"],
        ["Einheiten", _zahl(inputs.get("anzahl_ne")), "Zertifizierung", inputs.get("zertifizierung") or "-"],
    ]
    t = Table(grundlagen, colWidths=[27 * mm, 59 * mm, 27 * mm, 59 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{HELL}")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(f"#{DUNKEL}")),
        ("FONTNAME", (0, 0), (-1, -1), "GkSans"),
        ("FONTNAME", (0, 0), (0, -1), "GkSans-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "GkSans-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{LINIE}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story += [t, Spacer(1, 5 * mm)]

    unvollstaendig = bool(result.get("ist_unvollstaendig"))
    callout_text = (
        f"Schätzung unvollständig: {len(result.get('fehlende_positionen') or [])} Position(en) ohne Betrag."
        if unvollstaendig else "Schätzung vollständig: alle angezeigten Positionen besitzen einen Endbetrag."
    )
    callout_fg, callout_bg = (ROT, ROT_HELL) if unvollstaendig else (GRUEN, GRUEN_HELL)
    summary = Table([
        [Paragraph("Gesamtschätzung", styles["body"]), Paragraph(f"<b>{_chf(result.get('gesamt_betrag'))}</b>", styles["right"])],
        [Paragraph(callout_text, styles["body"]), ""],
    ], colWidths=[115 * mm, 57 * mm])
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{callout_bg}")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(f"#{callout_fg}")),
        ("SPAN", (0, 1), (1, 1)),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor(f"#{callout_fg}")),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story += [summary, Paragraph("Kostenzusammenstellung", styles["h1"])]

    for gruppe in result.get("gruppen") or []:
        daten = [["BKP", "Position", "Datenbasis / Quelle", "Betrag"]]
        for p in gruppe.get("positionen") or []:
            status = "Manuell" if p.get("quelle") == "manuell" else p.get("status_datenbasis") or "-"
            status = status.replace("–", "-").replace("—", "-")
            fg, _ = _status_farbe(p)
            daten.append([
                p.get("bkp_nr") or "", Paragraph(p.get("bezeichnung") or "", styles["body"]),
                Paragraph(f'<font color="#{fg}">{status}</font>', styles["small"]),
                Paragraph(_chf(p.get("betrag")), styles["right"]),
            ])
        daten.append(["", Paragraph(f"<b>Zwischentotal {gruppe.get('gruppe_nr')} {gruppe.get('name') or ''}</b>", styles["body"]), "",
                      Paragraph(f"<b>{_chf(gruppe.get('betrag'))}</b>", styles["right"])])
        table = Table(daten, colWidths=[17 * mm, 75 * mm, 43 * mm, 37 * mm], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{DUNKEL}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "GkSans-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7.5),
            ("FONTNAME", (0, 1), (0, -1), "GkSans"),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("ALIGN", (3, 1), (3, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (0, 1), (-1, -2), 0.3, colors.HexColor(f"#{LINIE}")),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(f"#{HELL}")),
            ("LINEABOVE", (0, -1), (-1, -1), 0.7, colors.HexColor(f"#{DUNKEL}")),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story += [Paragraph(f"BKP {gruppe.get('gruppe_nr')} - {gruppe.get('name') or ''}", styles["h1"]), table]

    herkunft = [(p, h) for p in _alle_positionen(result) for h in (p.get("herkunft") or [])]
    if herkunft:
        story += [PageBreak(), Paragraph("Herkunftsnachweis", styles["title"]),
                  Paragraph("Nur Referenzprojekte mit einer tatsächlich eingerechneten Kostenangabe sind aufgeführt.", styles["body"]),
                  Spacer(1, 3 * mm)]
        daten = [["BKP", "Referenzprojekt", "Kosten", "Kennwert", "Gewicht"]]
        for p, h in herkunft:
            daten.append([
                p.get("bkp_nr") or "", Paragraph(h.get("name") or "-", styles["small"]),
                _chf(h.get("kosten")), f"{_zahl(h.get('kennwert'), 2)} {p.get('einheit') or ''}",
                f"{_zahl((h.get('gewicht') or 0) * 100, 1)} %",
            ])
        table = Table(daten, colWidths=[18 * mm, 65 * mm, 31 * mm, 37 * mm, 21 * mm], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{DUNKEL}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "GkSans-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "GkSans"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor(f"#{LINIE}")),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)

    doc.build(story, onFirstPage=fuss, onLaterPages=fuss)
    return buf.getvalue()


def erzeuge_grobkostenschaetzung_excel(projekt_name: str, inputs: dict, result: dict,
                                       variante: str) -> bytes:
    """Erzeugt eine editierbare XLSX mit Formeln und separatem Herkunftsnachweis."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kostenschätzung"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"

    thin = Side(style="thin", color=LINIE)
    ws.merge_cells("A1:H1")
    ws["A1"] = "Grobkostenschätzung Heizung"
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=DUNKEL)
    ws["A2"] = "Projekt"
    ws["B2"] = projekt_name
    ws["D2"] = "Variante"
    ws["E2"] = variante.title()
    ws["G2"] = "Stand"
    ws["H2"] = date.today()
    ws["H2"].number_format = "dd.mm.yyyy"
    ws["A3"] = "Gesamtschätzung"
    ws["A3"].font = Font(bold=True, color=DUNKEL)
    ws["B3"].font = Font(size=14, bold=True, color=DUNKEL)
    ws["B3"].number_format = '#,##0 "CHF"'
    ws["D3"] = "Status"
    ws["E3"] = "Unvollständig" if result.get("ist_unvollstaendig") else "Vollständig"
    ws["E3"].fill = PatternFill("solid", fgColor=ROT_HELL if result.get("ist_unvollstaendig") else GRUEN_HELL)
    ws["E3"].font = Font(bold=True, color=ROT if result.get("ist_unvollstaendig") else GRUEN)

    grundlagen = [
        ("Nutzung", inputs.get("nutzung")), ("Projektart", inputs.get("projektart")),
        ("Wärmeerzeuger", ", ".join(inputs.get("waermeerzeuger") or [])),
        ("Wärmeabgabe", ", ".join(inputs.get("waermeabgabe") or [])),
        ("EBF [m²]", inputs.get("ebf_m2")), ("Heizleistung [kW]", inputs.get("leistung_kw")),
        ("Einheiten", inputs.get("anzahl_ne")), ("Zertifizierung", inputs.get("zertifizierung")),
    ]
    for index, (label, wert) in enumerate(grundlagen):
        row = 5 + index // 4
        col = 1 + (index % 4) * 2
        ws.cell(row, col, label).font = Font(size=9, bold=True, color=GRAU)
        ws.cell(row, col + 1, wert if wert not in (None, "") else "-")

    header_row = 9
    headers = ["BKP", "Position", "Kennwert", "Einheit", "Berechnet [CHF]", "Manuell [CHF]", "Endbetrag [CHF]", "Datenbasis / Quelle"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.fill = PatternFill("solid", fgColor=DUNKEL)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center")

    row = header_row + 1
    gruppen_zeilen = []
    for gruppe in result.get("gruppen") or []:
        gruppen_row = row
        gruppen_zeilen.append(gruppen_row)
        ws.cell(row, 1, gruppe.get("gruppe_nr"))
        ws.cell(row, 2, gruppe.get("name"))
        for col in range(1, 9):
            ws.cell(row, col).fill = PatternFill("solid", fgColor="EEF2F6")
            ws.cell(row, col).font = Font(bold=True, color=DUNKEL)
        row += 1
        positions_zeilen = []
        for p in gruppe.get("positionen") or []:
            positions_zeilen.append(row)
            ws.cell(row, 1, p.get("bkp_nr"))
            ws.cell(row, 2, p.get("bezeichnung"))
            ws.cell(row, 3, p.get("kennwert"))
            ws.cell(row, 4, p.get("einheit"))
            ws.cell(row, 5, p.get("berechneter_betrag"))
            ws.cell(row, 6, p.get("manueller_betrag"))
            ws.cell(row, 7, f'=IF(F{row}<>"",F{row},E{row})')
            status = "Manuell" if p.get("quelle") == "manuell" else p.get("status_datenbasis")
            ws.cell(row, 8, status)
            fg, bg = _status_farbe(p)
            ws.cell(row, 8).fill = PatternFill("solid", fgColor=bg)
            ws.cell(row, 8).font = Font(color=fg)
            for col in range(1, 9):
                ws.cell(row, col).border = Border(bottom=thin)
            row += 1
        if positions_zeilen:
            refs = ",".join(f"G{r}" for r in positions_zeilen)
            ws.cell(gruppen_row, 7, f"=SUM({refs})")
            ws.cell(gruppen_row, 7).number_format = '#,##0 "CHF"'

    if gruppen_zeilen:
        ws["B3"] = "=SUM(" + ",".join(f"G{r}" for r in gruppen_zeilen) + ")"
    else:
        ws["B3"] = 0
    ws.auto_filter.ref = f"A{header_row}:H{row - 1}"
    for col in (3, 5, 6, 7):
        for r in range(header_row + 1, row):
            ws.cell(r, col).number_format = '#,##0.00' if col == 3 else '#,##0 "CHF"'
            ws.cell(r, col).alignment = Alignment(horizontal="right")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 43
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 29
    ws.print_title_rows = f"1:{header_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.center.text = "Grobkostenschätzung - Seite &P von &N"

    ref_ws = wb.create_sheet("Referenzdetails")
    ref_ws.sheet_view.showGridLines = False
    ref_headers = ["BKP", "Position", "Referenzprojekt", "Datum", "EBF [m²]", "Leistung [kW]", "Kosten [CHF]", "Bezugsgrösse", "Kennwert", "Gewicht"]
    for col, header in enumerate(ref_headers, 1):
        cell = ref_ws.cell(1, col, header)
        cell.fill = PatternFill("solid", fgColor=DUNKEL)
        cell.font = Font(bold=True, color="FFFFFF")
    ref_row = 2
    for p in _alle_positionen(result):
        for h in p.get("herkunft") or []:
            werte = [
                p.get("bkp_nr"), p.get("bezeichnung"), h.get("name"), h.get("datum_abrechnung"),
                h.get("ebf_m2"), h.get("leistung_kw"), h.get("kosten"), h.get("treiber_wert"),
                h.get("kennwert"), h.get("gewicht"),
            ]
            for col, wert in enumerate(werte, 1):
                ref_ws.cell(ref_row, col, wert)
                ref_ws.cell(ref_row, col).border = Border(bottom=thin)
            ref_ws.cell(ref_row, 7).number_format = '#,##0 "CHF"'
            ref_ws.cell(ref_row, 9).number_format = "#,##0.00"
            ref_ws.cell(ref_row, 10).number_format = "0.0%"
            ref_row += 1
    ref_ws.freeze_panes = "A2"
    ref_ws.auto_filter.ref = f"A1:J{max(1, ref_row - 1)}"
    for index, width in enumerate([12, 40, 30, 13, 13, 16, 17, 17, 15, 12], 1):
        ref_ws.column_dimensions[get_column_letter(index)].width = width
    ref_ws.print_title_rows = "1:1"
    ref_ws.sheet_properties.pageSetUpPr.fitToPage = True
    ref_ws.page_setup.paperSize = ref_ws.PAPERSIZE_A4
    ref_ws.page_setup.orientation = ref_ws.ORIENTATION_LANDSCAPE
    ref_ws.page_setup.fitToWidth = 1
    ref_ws.page_setup.fitToHeight = 0
    ref_ws.oddFooter.center.text = "Referenzdetails - Seite &P von &N"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
