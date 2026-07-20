import io

from openpyxl import load_workbook
from pypdf import PdfReader

from app.export.grobkostenschaetzung import (
    erzeuge_grobkostenschaetzung_excel,
    erzeuge_grobkostenschaetzung_pdf,
)


INPUTS = {
    "nutzung": "MFH", "projektart": "Neubau", "waermeerzeuger": ["Erdsonden-WP"],
    "waermeabgabe": ["FBH"], "ebf_m2": 1700, "leistung_kw": 60,
    "anzahl_ne": 13, "zertifizierung": "Gesetz",
}

RESULT = {
    "gesamt_betrag": 122000,
    "ist_unvollstaendig": False,
    "fehlende_positionen": [],
    "gruppen": [
        {
            "gruppe_nr": "242", "name": "Wärmeerzeugung", "betrag": 90000,
            "positionen": [
                {
                    "bkp_nr": "242.3", "bezeichnung": "Wärmepumpe Sole/Wasser",
                    "kennwert": 1500, "einheit": "CHF/kW", "berechneter_betrag": 90000,
                    "manueller_betrag": None, "betrag": 90000, "quelle": "referenzen",
                    "status_datenbasis": "Begrenzte Datengrundlage",
                    "herkunft": [{
                        "name": "Referenz MFH", "datum_abrechnung": "2025-05-01",
                        "waermeerzeuger": ["Erdsonden-WP", "Gas"],
                        "ebf_m2": 1650, "leistung_kw": 58, "kosten": 87000,
                        "treiber_wert": 58, "kennwert": 1500, "gewicht": 0.91,
                    }],
                },
            ],
        },
        {
            "gruppe_nr": "243", "name": "Wärmeverteilung", "betrag": 32000,
            "positionen": [
                {
                    "bkp_nr": "243.3a", "bezeichnung": "Flächenheizung (Bodenheizung)",
                    "kennwert": None, "einheit": "CHF/m² EBF", "berechneter_betrag": None,
                    "manueller_betrag": 32000, "betrag": 32000, "quelle": "manuell",
                    "status_datenbasis": "Keine Angaben", "herkunft": [],
                },
            ],
        },
    ],
}


def test_pdf_export_ist_lesbar_und_enthaelt_kerndaten():
    pdf = erzeuge_grobkostenschaetzung_pdf("Projekt Test", INPUTS, RESULT, "netto")
    assert pdf.startswith(b"%PDF")
    reader = PdfReader(io.BytesIO(pdf))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Grobkostenschätzung Heizung" in text
    assert "242.3" in text
    assert "CHF 122'000" in text
    assert "Herkunftsnachweis" in text


def test_excel_export_hat_formeln_formatierung_und_referenzdetails():
    xlsx = erzeuge_grobkostenschaetzung_excel("Projekt Test", INPUTS, RESULT, "netto")
    wb = load_workbook(io.BytesIO(xlsx), data_only=False)
    assert wb.sheetnames == ["Kostenschätzung", "Referenzdetails", "Manuelle Werte"]
    ws = wb["Kostenschätzung"]
    assert ws["A1"].value == "Grobkostenschätzung Heizung"
    assert str(ws["B3"].value).startswith("=SUM(")
    assert ws.freeze_panes == "A10"
    manuell_row = next(row for row in range(10, ws.max_row + 1) if ws.cell(row, 1).value == "243.3a")
    assert ws.cell(manuell_row, 6).value == 32000
    assert ws.cell(manuell_row, 7).value == f'=IF(F{manuell_row}<>"",F{manuell_row},E{manuell_row})'
    assert ws.cell(manuell_row, 8).value == "Manuell"
    refs = wb["Referenzdetails"]
    assert refs["C2"].value == "Referenz MFH"
    assert refs["D2"].value == "Erdsonden-WP + Gas"
    assert refs["H2"].value == 87000


def test_unvollstaendiger_export_bezeichnet_summe_als_teilbetrag():
    result = {**RESULT, "ist_unvollstaendig": True, "fehlende_positionen": ["243.1"]}
    xlsx = erzeuge_grobkostenschaetzung_excel("Projekt Test", INPUTS, result, "netto")
    wb = load_workbook(io.BytesIO(xlsx), data_only=False)
    assert wb["Kostenschätzung"]["A3"].value == "Teilbetrag bekannte Positionen"

    pdf = erzeuge_grobkostenschaetzung_pdf("Projekt Test", INPUTS, result, "netto")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf)).pages)
    assert "Teilbetrag bekannte Positionen" in text


def test_freigegebener_status_erscheint_in_pdf_und_excel():
    inputs = {**INPUTS, "_schaetzung_status": "freigegeben", "_freigegeben_at": "2026-07-20T12:00:00"}
    xlsx = erzeuge_grobkostenschaetzung_excel("Projekt Test", inputs, RESULT, "netto")
    wb = load_workbook(io.BytesIO(xlsx), data_only=False)
    assert wb["Kostenschätzung"]["H3"].value == "Freigegeben"

    pdf = erzeuge_grobkostenschaetzung_pdf("Projekt Test", inputs, RESULT, "netto")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf)).pages)
    assert "Bearbeitungsstatus: Freigegeben" in text


def test_manuelle_dokumentation_erscheint_in_pdf_und_excel():
    inputs = {**INPUTS, "manuelle_notizen": {"netto": {"243.3a": {
        "begruendung": "Richtofferte Unternehmer", "quelle": "Offerte vom 18.07.2026",
        "bearbeiter": "Dominic", "geaendert_at": "2026-07-20T12:00:00",
    }}}}
    xlsx = erzeuge_grobkostenschaetzung_excel("Projekt Test", inputs, RESULT, "netto")
    wb = load_workbook(io.BytesIO(xlsx), data_only=False)
    manuell = wb["Manuelle Werte"]
    assert manuell["A2"].value == "243.3a"
    assert manuell["D2"].value == "Richtofferte Unternehmer"
    assert manuell["F2"].value == "Dominic"

    pdf = erzeuge_grobkostenschaetzung_pdf("Projekt Test", inputs, RESULT, "netto")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf)).pages)
    assert "Dokumentation manueller Werte" in text
    assert "Richtofferte Unternehmer" in text
