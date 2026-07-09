"""Automatischer Abruf des Schweizer Baupreisindex (BFS) über opendata.swiss.

Verifiziert (2026-07-09, live getestet): die genaue Datensatz-ID auf
opendata.swiss ändert sich je nach BFS-Aktualisierung/Sprachversion (z.B.
"...-u1" / "...-u2" / "...-u3-1") — darum Suche per `package_search` statt
eine feste ID (bricht nicht bei der nächsten Umbenennung). Die Rohdaten
liegen als Excel-Datei vor (XLSX, mehrere Tabellenblätter "Basis <Jahr> = 100"),
nicht als CSV. Kennwert: Region "Schweiz", Kategorie "Baugewerbe : Total"
(nationaler Gesamtindex, halbjährlich April/Oktober).
"""
import io
from datetime import date
from typing import Optional

CKAN_SEARCH = "https://ckan.opendata.swiss/api/3/action/package_search"
SUCHBEGRIFF = "Baupreisindex"
_HEADERS = {"User-Agent": "Mozilla/5.0 (Heizungscockpit/1.0; +https://energienachweise.com)"}
_MONAT = {"Oktober": 10, "April": 4}


def _finde_xlsx_url(client) -> Optional[str]:
    resp = client.get(CKAN_SEARCH, params={"q": SUCHBEGRIFF, "rows": 5})
    resp.raise_for_status()
    payload = resp.json()
    if not payload.get("success"):
        return None
    for result in payload.get("result", {}).get("results", []):
        for resource in result.get("resources", []):
            if "xls" in (resource.get("format") or "").lower():
                return resource.get("url") or resource.get("download_url")
    return None


def _lies_schweiz_gesamtindex(xlsx_bytes: bytes) -> list:
    """Liest Periode+Wert für Region 'Schweiz' / Kategorie 'Baugewerbe : Total'.

    Struktur (mehrfach live geprüft): eines der numerisch benannten
    Tabellenblätter ("1998"/"2010"/.../"2020" = Basisjahr) enthält eine Zeile
    mit den Monatsnamen ("Oktober"/"April") — direkt darunter die Jahreszahlen,
    je eine Spalte ab Spalte D pro Erhebungsperiode. "Schweiz" markiert den
    Beginn des nationalen Blocks; die Zeile direkt danach ist der Gesamtindex.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    basis_sheets = sorted((s for s in wb.sheetnames if s.isdigit()), key=int)
    if not basis_sheets:
        return []
    ws = wb[basis_sheets[0]]

    monat_zeile = jahr_zeile = None
    for r in range(1, 15):
        werte = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(w in _MONAT for w in werte):
            monat_zeile, jahr_zeile = r, r + 1
            break
    if monat_zeile is None:
        return []

    schweiz_zeile = None
    for r in range(monat_zeile, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "Schweiz":
            schweiz_zeile = r + 1
            break
    if schweiz_zeile is None:
        return []

    eintraege = []
    for c in range(4, ws.max_column + 1):
        monat = ws.cell(row=monat_zeile, column=c).value
        jahr = ws.cell(row=jahr_zeile, column=c).value
        wert = ws.cell(row=schweiz_zeile, column=c).value
        if monat in _MONAT and isinstance(jahr, (int, float)) and isinstance(wert, (int, float)):
            eintraege.append({"periode": date(int(jahr), _MONAT[monat], 1), "wert": float(wert)})
    return eintraege


def fetch_bfs_baupreisindex() -> dict:
    """Bestmöglicher Abruf. Gibt immer {"erfolg", "meldung", "eintraege"} zurück,
    wirft nie ungefangen. eintraege: Liste von {"periode": date, "wert": float}."""
    try:
        import httpx
    except ImportError:
        return {"erfolg": False, "meldung": "httpx ist nicht installiert.", "eintraege": []}

    try:
        with httpx.Client(timeout=20, headers=_HEADERS, follow_redirects=True) as client:
            xlsx_url = _finde_xlsx_url(client)
            if not xlsx_url:
                return {"erfolg": False, "meldung": "Keine Excel-Ressource beim BFS-Datensatz gefunden.", "eintraege": []}
            resp = client.get(xlsx_url)
            resp.raise_for_status()
            xlsx_bytes = resp.content
    except Exception as e:
        return {"erfolg": False, "meldung": f"Abruf fehlgeschlagen: {e}", "eintraege": []}

    try:
        eintraege = _lies_schweiz_gesamtindex(xlsx_bytes)
    except Exception as e:
        return {"erfolg": False, "meldung": f"Excel-Datei konnte nicht gelesen werden: {e}", "eintraege": []}

    if not eintraege:
        return {"erfolg": False, "meldung": "Keine auswertbaren Indexwerte gefunden (Struktur unerwartet).", "eintraege": []}
    return {
        "erfolg": True,
        "meldung": f"{len(eintraege)} Indexwerte gefunden (Schweiz, Baugewerbe Total, {eintraege[0]['periode'].year}–{eintraege[-1]['periode'].year}).",
        "eintraege": eintraege,
    }
